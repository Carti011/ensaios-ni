import pytest

from ensaios_ni.dominio.serie import SerieTemporal
from ensaios_ni.persistencia.exportadores import exportar_csv_excel_br, exportar_xlsx


def test_csv_excel_br_usa_ponto_e_virgula_e_decimal_virgula(tmp_path):
    caminho = tmp_path / "ensaio.csv"
    serie = SerieTemporal(
        canais=["Mod1/ai0", "Mod3/ai0"],
        unidades={"Mod1/ai0": "kgf", "Mod3/ai0": "ue"},
        taxa_hz=50.0,
        dados={"Mod1/ai0": [200.0, 201.5], "Mod3/ai0": [10.0, 11.25]},
    )

    exportar_csv_excel_br(serie, caminho)

    assert caminho.read_text(encoding="utf-8-sig") == (
        "tempo_s;Mod1/ai0 (kgf);Mod3/ai0 (ue)\n"
        "0,0;200,0;10,0\n"
        "0,02;201,5;11,25\n"
    )


def test_csv_excel_br_seleciona_sinais_na_ordem_da_serie(tmp_path):
    caminho = tmp_path / "ensaio.csv"
    serie = SerieTemporal(
        canais=["Mod1/ai0", "Mod1/ai1", "Mod3/ai0"],
        unidades={"Mod1/ai0": "kgf", "Mod1/ai1": "bar", "Mod3/ai0": "ue"},
        taxa_hz=50.0,
        dados={"Mod1/ai0": [200.0], "Mod1/ai1": [3.0], "Mod3/ai0": [10.0]},
    )

    exportar_csv_excel_br(serie, caminho, sinais=["Mod3/ai0", "Mod1/ai0"])

    assert caminho.read_text(encoding="utf-8-sig") == (
        "tempo_s;Mod1/ai0 (kgf);Mod3/ai0 (ue)\n"
        "0,0;200,0;10,0\n"
    )


def test_csv_excel_br_recorta_pela_janela_de_tempo(tmp_path):
    # tempos 0; 0,02; 0,04; 0,06 — a janela [0,02; 0,04] mantém o tempo absoluto
    caminho = tmp_path / "ensaio.csv"
    serie = SerieTemporal(
        canais=["Mod1/ai0"],
        unidades={"Mod1/ai0": "kgf"},
        taxa_hz=50.0,
        dados={"Mod1/ai0": [10.0, 11.0, 12.0, 13.0]},
    )

    exportar_csv_excel_br(serie, caminho, inicio_s=0.02, fim_s=0.04)

    assert caminho.read_text(encoding="utf-8-sig") == (
        "tempo_s;Mod1/ai0 (kgf)\n"
        "0,02;11,0\n"
        "0,04;12,0\n"
    )


def test_csv_excel_br_recusa_sinal_inexistente(tmp_path):
    caminho = tmp_path / "ensaio.csv"
    serie = SerieTemporal(
        canais=["Mod1/ai0"],
        unidades={"Mod1/ai0": "kgf"},
        taxa_hz=50.0,
        dados={"Mod1/ai0": [200.0]},
    )

    with pytest.raises(ValueError, match="inexistentes"):
        exportar_csv_excel_br(serie, caminho, sinais=["Mod9/ai9"])


def test_xlsx_grava_cabecalho_e_numeros_nativos(tmp_path):
    import openpyxl

    caminho = tmp_path / "ensaio.xlsx"
    serie = SerieTemporal(
        canais=["Mod1/ai0", "Mod3/ai0"],
        unidades={"Mod1/ai0": "kgf", "Mod3/ai0": "ue"},
        taxa_hz=50.0,
        dados={"Mod1/ai0": [200.0, 201.5], "Mod3/ai0": [10.0, 11.25]},
    )

    exportar_xlsx(serie, caminho)

    planilha = openpyxl.load_workbook(caminho).active
    linhas = list(planilha.iter_rows(values_only=True))
    assert linhas == [
        ("tempo_s", "Mod1/ai0 (kgf)", "Mod3/ai0 (ue)"),
        (0.0, 200.0, 10.0),
        (0.02, 201.5, 11.25),
    ]


def test_xlsx_recorta_pela_janela_de_tempo(tmp_path):
    import openpyxl

    caminho = tmp_path / "ensaio.xlsx"
    serie = SerieTemporal(
        canais=["Mod1/ai0"],
        unidades={"Mod1/ai0": "kgf"},
        taxa_hz=50.0,
        dados={"Mod1/ai0": [10.0, 11.0, 12.0, 13.0]},
    )

    exportar_xlsx(serie, caminho, inicio_s=0.04)

    aba = openpyxl.load_workbook(caminho).active
    assert list(aba.iter_rows(values_only=True)) == [
        ("tempo_s", "Mod1/ai0 (kgf)"),
        (0.04, 12.0),
        (0.06, 13.0),
    ]


def test_janela_invertida_e_recusada_sem_criar_arquivo(tmp_path):
    caminho = tmp_path / "ensaio.csv"
    serie = SerieTemporal(
        canais=["Mod1/ai0"], unidades={}, taxa_hz=50.0, dados={"Mod1/ai0": [1.0, 2.0, 3.0]}
    )

    with pytest.raises(ValueError, match="inicio_s.*fim_s|janela"):
        exportar_csv_excel_br(serie, caminho, inicio_s=0.04, fim_s=0.02)
    assert not caminho.exists()


def test_janela_sem_amostras_e_recusada(tmp_path):
    caminho = tmp_path / "ensaio.csv"
    serie = SerieTemporal(
        canais=["Mod1/ai0"], unidades={}, taxa_hz=50.0, dados={"Mod1/ai0": [1.0, 2.0, 3.0]}
    )

    with pytest.raises(ValueError, match="não contém amostras"):
        exportar_csv_excel_br(serie, caminho, inicio_s=10.0)
    assert not caminho.exists()


def test_xlsx_seleciona_sinais(tmp_path):
    import openpyxl

    caminho = tmp_path / "ensaio.xlsx"
    serie = SerieTemporal(
        canais=["Mod1/ai0", "Mod3/ai0"],
        unidades={"Mod1/ai0": "kgf", "Mod3/ai0": "ue"},
        taxa_hz=50.0,
        dados={"Mod1/ai0": [200.0], "Mod3/ai0": [10.0]},
    )

    exportar_xlsx(serie, caminho, sinais=["Mod1/ai0"])

    planilha = openpyxl.load_workbook(caminho).active
    linhas = list(planilha.iter_rows(values_only=True))
    assert linhas == [
        ("tempo_s", "Mod1/ai0 (kgf)"),
        (0.0, 200.0),
    ]
